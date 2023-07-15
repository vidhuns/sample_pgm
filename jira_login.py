from jira import JIRA
import ssl
from openpyxl import Workbook
import smtplib
import email
from email import encoders
from email.message import EmailMessage
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

#import base64

#jira = JIRA(basic_auth=('vsahayaraj.contractor@libertyglobal.com','Mjk4MzExODc0ODk4On0ZPDQBm1Vi4NZpVKUPocgx3SET'), options={"server": "https://jira.lgi.io"} )
jira = JIRA(basic_auth=('vsahayaraj.contractor','Liberty13@20505$#'), options={"server": "https://jira.lgi.io"} )

mail_alias = {'vidhun': 'vsahayaraj.contractor@libertyglobal.com', 'praveen': 'psrinivas.ext@libertyglobal.com'}
key_list = []
summary_list = []
print("enter the JQL query you want")
jql_query = input("")
#issues_in_project = jira.search_issues('project = MVXREQ and assignee = vsahayaraj.contractor')
issues_in_project = jira.search_issues(str(jql_query))
for issue in issues_in_project:
    key_list.append(issue.key)
    summary_list.append(issue.fields.summary)

wb = Workbook()
ws = wb.active
key_row = 1
summary_row = 1
start_column = 1
for key in key_list:
    ws.cell(row=key_row, column=start_column).value = key
    key_row += 1
for summary in summary_list:
    ws.cell(row=summary_row, column=start_column+1).value = summary
    summary_row += 1

location = str(r"C:\Users\vsahayaraj\Vidhun\Personal\Learning\python\\")
print("Enter the file name to save the data")
filename = str(input(""))
filename_ext = filename + str('.xlsx')
path = location + filename_ext
#print(path)
wb.save(path)
print("The report is saved in the", path) 
sender = 'vsahayaraj.contractor@libertyglobal.com'
print("enter the receipent name")
receiver_name = str(input(""))

def mail_code():        
 sender = 'vsahayaraj.contractor@libertyglobal.com'
 body = "This is an email with attachment sent from Python"
 subject = 'test email for report'    
# message = EmailMessage()
 message = MIMEMultipart()
 message["From"] = sender
 message["To"] = receiver
 message["Subject"] = subject
 message.attach(MIMEText(body, "plain"))
 filename = path
 with open(filename, "rb") as attachment:
     part = MIMEBase("application", "octet-stream")
     part.set_payload(attachment.read())

 encoders.encode_base64(part)
 part.add_header(
    "Content-Disposition",
    f"attachment; filename= {filename_ext}",
 )
 message.attach(part)
 text = message.as_string()
 #print(text)
 #print("text is before try part")
 try:
   smtpObj = smtplib.SMTP('172.30.180.9')
   smtpObj.sendmail(sender, receiver, text)
   #smtpObj.send_message(text)         
   print ("Successfully sent email")
 except:
   print ("Error: unable to send email")

if receiver_name in mail_alias:
     receiver = mail_alias[receiver_name] 
     print(receiver)
     mail_code()

else:
    print("the receiver name is not in the list\n, Do you want to add? then presss (yes/no)")
    name_response= str(input())
    if(name_response== 'yes'):
        print("enter the receipent name")
        receiver_name_1 = str(input(""))  
        if receiver_name_1 in mail_alias:
         receiver_1 = mail_alias[receiver_name_1] 
        print("inside elif")
    else:
        print("sorry stopping the program")
#context = ssl.create_default_context()
#with smtplib.SMTP_SSL("172.30.180.9, context=context") as server:  
    


  



